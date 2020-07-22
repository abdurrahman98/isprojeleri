import requests as req
from bs4 import BeautifulSoup
from openpyxl import workbook


def urunBilgileri(request):
    try:
        web_get = req.get(request.POST["web"])
        web_icerik = BeautifulSoup(web_get.content, "html.parser")
        urunAdi = web_icerik.find("h1", attrs={"id": "product-name"})
        fiyat = web_icerik.find("span", attrs={
            "data-bind": "text: product().currentListing.currentPriceBeforePoint + ',' + product().currentListing.currentPriceAfterPoint"})
        veri = web_icerik.find("div", attrs={"id": "productTechSpecContainer"})
        veriler = dict()

        book = workbook.Workbook()
        sayfa = book.create_sheet("sayfa", 1)
        sayfa.cell(row=1, column=1, value=f"Ürün Adı")
        sayfa.cell(row=2, column=1, value=f"{urunAdi.text.strip()}")
        sayfa.cell(row=1, column=2, value=f"Fiyat")
        sayfa.cell(row=2, column=2, value=f"{fiyat.text}")
        print(veri.find_all("table"))
        for y in veri.find_all("table"):

            for index, z in enumerate(y.find_all("th")):
                veriler[f"{z.text}"] = y.find_all("td")[index].text
                sayfa.cell(row=1, column=index + 3, value=f"{z.text}")
                print("Başlık", z.text, sep=": ", end="")
                print(" İçerik", y.find_all('td')[index - 1].text, sep=": ", end="\n")
                sayfa.cell(row=2, column=index + 3, value=f"{y.find_all('td')[index].text}")

        book.save("deneme.xlsx")
        book.close()
    except:
        pass


def urunBilgileriAll(urunAll):
    book = workbook.Workbook()
    sayfa = book.create_sheet("sayfa", 1)
    for sayac,i in enumerate(urunAll,-1):
        sayac += 2
        web_get = req.get("https://www.hepsiburada.com/" + str(i))
        web_icerik = BeautifulSoup(web_get.content, "html.parser")
        urunAdi = web_icerik.find("h1", attrs={"id": "product-name"})
        fiyat = web_icerik.find("span", attrs={
            "data-bind": "text: product().currentListing.currentPriceBeforePoint + ',' + product().currentListing.currentPriceAfterPoint"})
        veri = web_icerik.find("div", attrs={"id": "productTechSpecContainer"})
        veriler = dict()


        sayfa.cell(row=sayac, column=1, value=f"Ürün Adı")
        sayfa.cell(row=sayac+1, column=1, value=f"{urunAdi.text.strip()}")
        sayfa.cell(row=sayac, column=2, value=f"Fiyat")
        sayfa.cell(row=sayac+1, column=2, value=f"{fiyat.text}")
        print(veri.find_all("table"))
        for y in veri.find_all("table"):

            for index, z in enumerate(y.find_all("th")):
                veriler[f"{z.text}"] = y.find_all("td")[index].text
                sayfa.cell(row=sayac, column=index + 3, value=f"{z.text}")

                sayfa.cell(row=sayac+1, column=index + 3, value=f"{y.find_all('td')[index].text}")

    book.save("deneme.xlsx")
    book.close()





def urunBilgileriHepsi(urunAll):
    book = workbook.Workbook()
    sayfa = book.create_sheet("sayfa", 1)
    sayfa.cell(row=1,column=1,value="Ürün adı")
    sayfa.cell(row=1,column=2,value="Fiyat")
    sayfa.cell(row=1,column=3,value="Stok Kodu")
    sayfa.cell(row=1,column=4,value="image url 1")
    sayfa.cell(row=1,column=5,value="image url 2")
    sayfa.cell(row=1,column=6,value="image url 3")
    sayfa.cell(row=1,column=7,value="image url 4")
    sayfa.cell(row=1,column=8,value="image url 5")
    sayfa.cell(row=1,column=9,value="image url 6")
    sayfa.cell(row=1,column=10,value="image url 7")
    say=2

    for i in urunAll:
        web_get = req.get("https://www.hepsiburada.com" + str(i))
        web_icerik = BeautifulSoup(web_get.content, "html.parser")

        urunAdi = web_icerik.find("h1", attrs={"id": "product-name"})
        fiyat = web_icerik.find("span", attrs={
            "data-bind": "text: product().currentListing.currentPriceBeforePoint + ',' + product().currentListing.currentPriceAfterPoint"})
        veri = web_icerik.find("div", attrs={"id": "productTechSpecContainer"})
        sayfa.cell(row=say,column=1,value=f"{urunAdi.text.lstrip().rstrip()}")
        sayfa.cell(row=say,column=2,value=f"{fiyat.text.lstrip().rstrip()}")
        for y in veri.find_all("table"):

            for index, z in enumerate(y.find_all("th")):
                if z.text=="Stok Kodu":
                    sayfa.cell(row=say,column=3,value=f"{y.find_all('td')[index].text}")


                    break
        print(web_icerik.find("div", attrs={"id": "productDetailsCarousel"}))
        # for k in range(0,100,1):
        #
        #     print(web_icerik.find("div", attrs={"id": "productDetailsCarousel"}).find_all("img")[k])


        for index,y in enumerate(web_icerik.find("div",attrs={"id":"productDetailsCarousel"}).find_all("img",attrs={"itemprop":"image"}),4):
            try:
                sayfa.cell(row=say,column=index,value=f"{y['src']}")
            except:
                sayfa.cell(row=say,column=index,value=f"{y['data-src']}")
            # print(str(say)+str(y))


        # for y in web_icerik.find_all("div",attrs={"id":"productDetailsCarousel"}):
        #
        #     for z in y.find_all("div",attrs={"class":"owl-item"}):
        #         print(z.find("img")["src"])
        say += 1
    book.save("deneme.xlsx")
    book.close()

        # for y in web_icerik.find_all("div",attrs={"style":"transform:translate3d(-1544px,0px,0px);transition: all 0.25s ease 0s;width: 4632px;"}):
        #     print(y)

def aramaBilgileri(request):
    urunAll = list()
    say = 2
    ilksayfa = req.get("https://www.hepsiburada.com/ara?q=" + str(request.POST["arama"]).strip().replace(" ", "+"))
    ilksayfa = BeautifulSoup(ilksayfa.content, "html.parser")
    urunler = ilksayfa.find_all("li", attrs={"class": "search-item col lg-1 md-1 sm-1 custom-hover not-fashion-flex"})
    for i in urunler:
        urunAll.append(i.find("a")["href"])
    # print(ilksayfa.find_all("li",attrs={"class":"search-item col lg-1 md-1 sm-1 custom-hover not-fashion-flex"}))
    while True:

        web_get = req.get("https://www.hepsiburada.com/ara?q=" +
                          str(request.POST["arama"]).strip().replace(" ", "+") +
                          "&sayfa=" + str(say))

        web_get = BeautifulSoup(web_get.content, "html.parser")

        if ilksayfa.text == web_get.text:

            break
        urunler = web_get.find_all("li",
                                   attrs={"class": "search-item col lg-1 md-1 sm-1 custom-hover not-fashion-flex"})

        for i in urunler:

            urunAll.append(i.find("a")["href"])

            # urunAll.append(i.get("href"))
            # print(i.find("a")["href"])

        say += 1
    urunBilgileriHepsi(urunAll)
